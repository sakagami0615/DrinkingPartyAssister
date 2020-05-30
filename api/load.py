import json
import openpyxl


def LoadJsonFile(filepath):
	
	with open(filepath, 'r', encoding='utf-8') as json_file:
		json_dict = json.load(json_file)
	return json_dict


def LoadMemberStatus(filepath, party_name):

	INIT_ROW = 2
	INIT_COL = 4
	HEADER_ITEMS = ['名前', '階級', '出欠', '会費']

	sheet_name = '{}(会費管理)'.format(party_name)
 
	workbook = openpyxl.load_workbook(filepath, data_only=True)
	sheet = workbook[sheet_name]
	
	row = INIT_ROW

	member_dict = {'Members': [], 'Num': None}

	while True:
		if sheet.cell(row, INIT_COL).value != None :
			
			status_dict = {}
			for (dcol, item) in enumerate(HEADER_ITEMS):
				
				if item == '名前' :
					status_dict['Name'] = sheet.cell(row, INIT_COL + dcol).value
				elif item == '階級' :
					status_dict['Rank'] = sheet.cell(row, INIT_COL + dcol).value
				elif item == '出欠' :
					status_dict['Attend'] = sheet.cell(row, INIT_COL + dcol).value
				elif item == '会費' :
					pass
				
				member_dict['Members'].append(status_dict)
			
			row = row + 1

		else:
			break
	
	member_dict['Num'] = row - INIT_ROW

	return member_dict
import re
import time
import openpyxl

from bs4 import BeautifulSoup
from selenium import webdriver


import load


"""
Selenium IE_Driver 仕様の準備
https://qiita.com/tsuttie/items/372f5d4cad37650711f1

https://qiita.com/__init__/items/1a90bfbbd2b1015bb90d

"""

SYSTEM_PATAM_FILEPATH = 'system_param.json'

SEARCH_PARAM_FILEPATH = '../お店探索条件.json'
INPUT_EXCEL_FILEPATH = '../DrinkingPartySheet.xlsm'



class Assister:

	def __init__(self):
		
		self.search_param = load.LoadJsonFile(SEARCH_PARAM_FILEPATH)
		self.system_param = load.LoadJsonFile(SYSTEM_PATAM_FILEPATH)

		self.member_dict = load.LoadMemberStatus(INPUT_EXCEL_FILEPATH, self.search_param['PartyName'])

		self.driver = webdriver.Ie(self.system_param['3rdParty']['IeDriver'])

	
	def Close(self):
		
		self.driver.close()
	

	def LoadWebPageTop(self):
		
		self.driver.get(self.system_param['HotPepper']['Url'])
		time.sleep(1)
	

	def SetShopConditions(self):
		
		area_index = self.system_param['HotPepper']['Area'][self.search_param['Area']]
		prefecture_index = self.system_param['HotPepper']['Prefecture'][self.search_param['Prefecture']]
		self.driver.find_element_by_xpath(self.system_param['HotPepper']['Json']['PrefectureButton'].format(area_index, prefecture_index)).click()
		time.sleep(1)

		self.driver.find_element_by_xpath(self.system_param['HotPepper']['Json']['CityInput']).send_keys(self.search_param['City'])
		self.driver.find_element_by_xpath(self.system_param['HotPepper']['Json']['CityButton']).click()
		time.sleep(1)
	

	# 店の名前とURLを取得
	def GetShopDetail(self):

		def GetShopDetailTop(find_soup):
			shop_details = []
			
			for soup in find_soup:
				
				name = re.sub(r'[\u3000]', ' ', soup.find('h3', {'class': 'shopDetailStoreName'}).find('a').get_text())
				url = '{}{}'.format(self.system_param['HotPepper']['BaseURL'], soup.find('h3', {'class': 'shopDetailStoreName'}).find('a').get('href'))
				seat_search = re.match(r'[\t\n\r\f\v]*(\d+)', str(soup.find('li', {'class': 'shopDetailInfoCapacity'}).get_text()))

				seat_num = None if seat_search == None else int(seat_search[0])
				shop_details.append({
					'Name': name,
					'SeatNum': seat_num,
					'Url': url
				})
			return shop_details


		shop_details = []
		page_index = 0

		while True:
			page_soup = BeautifulSoup(self.driver.page_source, 'html5lib')
			curr_shop_details = []		
			curr_shop_details.extend(GetShopDetailTop(page_soup.find_all('div', {'class': 'shopDetailTop PR shopDetailWithCourseCalendar'})))
			curr_shop_details.extend(GetShopDetailTop(page_soup.find_all('div', {'class': 'shopDetailTop shopDetailWithCourseCalendar'})))
			
			for curr_shop_detail in curr_shop_details:
				if (curr_shop_detail['SeatNum'] != None) and (curr_shop_detail['SeatNum'] >= self.member_dict['Num']): 
					shop_details.append(curr_shop_detail)

			if len(shop_details) >= self.search_param['SearchNum']:
				shop_details = shop_details[:self.search_param['SearchNum']]
				break

			return_button_index = 10 if page_index == 0 else 11
			self.driver.find_element_by_xpath(self.system_param['HotPepper']['Json']['NextButton'].format(len(curr_shop_details) + 2, return_button_index)).click()

			time.sleep(1)

			page_index = page_index + 1
		
		return shop_details

	
	# 店のコース情報を取得
	def GetCourseDetail(self, shop_details):
		
		course_details = []

		for shop_detail in shop_details:

			self.driver.get('{}course/'.format(shop_detail['Url']))
			time.sleep(1)

			page_soup = BeautifulSoup(self.driver.page_source, 'html5lib')		
			course_soups = page_soup.find_all('div', {'class': 'courseCassette'})

			for course_soup in course_soups:
				
				name = re.sub(r'[\u3000]', ' ', course_soup.find('p', {'class': 'courseCassetteTitle'}).find('a').get_text())
				url = '{}{}'.format(shop_detail['Url'], course_soup.find('p', {'class': 'courseCassetteTitle'}).find('a').get('href'))
				money_soup = course_soup.find('span', {'class': 'priceNumber'})
				if (money_soup == None): continue

				money_str = money_soup.get_text().replace(',', '')
				if not money_str.isdigit(): continue
				money = int(money_str)
				
				if (money > min(self.search_param['MoneyRange'])) and(money <= max(self.search_param['MoneyRange'])):
					course_details.append({
						'ShopName': shop_detail['Name'],
						'SeatNum': shop_detail['SeatNum'],
						'ShopUrl': shop_detail['Url'],
						'CourseName': name,
						'CourseUrl': url,
						'Money': money
					})

		return course_details

	
	def AddResultSheet4Excel(self, course_details):

		SELECT_KEYS  = ['ShopName', 'CourseName', 'Money', 'SeatNum', 'CourseUrl']
		HEADER_NAMES = ['店名', 'コース名', '金額', 'シート数', 'URL']
		HEADER_LINE_NUM = 1
		
		workbook = openpyxl.load_workbook(INPUT_EXCEL_FILEPATH)
		tgt_sheet = workbook[self.member_dict['SheetName']]

		add_sheet_idx = workbook.index(tgt_sheet) + 1
		add_sheet_name = '{}(開催場所候補)'.format(self.search_param['PartyName'])
		
		sheet = workbook.create_sheet(index=add_sheet_idx, title=add_sheet_name)
		
		for (idx, name) in enumerate(HEADER_NAMES):			
			sheet.cell(HEADER_LINE_NUM, idx + 1).value = name
		
		for (row, course_detail) in enumerate(course_details):
			for (col, key) in enumerate(SELECT_KEYS):		
				sheet.cell(row + 1 + HEADER_LINE_NUM, col + 1).value = course_detail[key]
